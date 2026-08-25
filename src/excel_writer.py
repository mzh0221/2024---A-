# -*- coding: utf-8 -*-
"""
excel_writer.py —— 输出模块（模块13，§11.3）
================================================
用 openpyxl 按题目附件模板格式写入 result1.xlsx ~ result5.xlsx，结果保留 6 位小数。

实现方式：加载附件中的模板文件（含表头/行标签），填入计算值后另存到 answer 目录，
        既保证格式与模板完全一致，又不破坏原始附件。

模板结构（由附件 result1/result2/result4.xlsx 校验）：
  result1/result4：含“位置”“速度”两个工作表
      位置表：A 列为把手标签（龙头x、龙头y、第1节龙身x、…、龙尾x、龙尾（后）y），
              B 列起为各时刻（0 s,1 s,…  或  −100 s,−99 s,…）
              行 2+2h = 把手 h 的 x，行 3+2h = 把手 h 的 y（h=0..223）
      速度表：行 2+h = 把手 h 的速度（h=0..223），B 列起为各时刻
      把手标签：h=0→龙头；h=1..221→第h节龙身；h=222→龙尾；h=223→龙尾（后）
  result2：单工作表 Sheet1，A=标签，B=横坐标x，C=纵坐标y，D=速度（单时刻快照）
  result3/result5：题目未给模板，按数值结果生成简易表（问题3=最小螺距，问题5=最大速度）
"""

import os
import openpyxl

# ---------------- 路径 ----------------
# 本文件位于 answer/ 下；附件在同级根目录的 2024---A--main/A题 copy/附件/
ANSWER_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(ANSWER_DIR)                         # .../Downloads/2024---A--main
ATTACH_DIR = os.path.join(ROOT_DIR, "2024---A--main", "A题 copy", "附件")


def _round6(v):
    """保留 6 位小数（返回 float，openpyxl 写入即按数值格式）。"""
    return round(float(v), 6)


def write_result1_or_4(template_name, out_name, times, pos_by_t, vel_by_t):
    """写 result1.xlsx / result4.xlsx（位置+速度两表，结构相同）。
    times       : 时刻列表（result1 为 0..300；result4 为 −100..100）
    pos_by_t[k] : 第 k 个时刻的把手坐标 (N_HANDLE, 2)
    vel_by_t[k] : 第 k 个时刻的把手速度 (N_HANDLE,)
    """
    from common import N_HANDLE

    tpl = os.path.join(ATTACH_DIR, template_name)
    out = os.path.join(ANSWER_DIR, out_name)
    wb = openpyxl.load_workbook(tpl)

    # ---- 位置表 ----
    ws = wb["位置"]
    for c, t in enumerate(times):
        col = c + 2                       # B 列起
        xy = pos_by_t[c]
        for h in range(N_HANDLE):
            ws.cell(row=2 + 2 * h, column=col, value=_round6(xy[h, 0]))   # x
            ws.cell(row=3 + 2 * h, column=col, value=_round6(xy[h, 1]))   # y
    # ---- 速度表 ----
    ws = wb["速度"]
    for c, t in enumerate(times):
        col = c + 2
        v = vel_by_t[c]
        for h in range(N_HANDLE):
            ws.cell(row=2 + h, column=col, value=_round6(v[h]))
    wb.save(out)
    print("  已保存: %s" % out)
    return out


def write_result2(template_name, out_name, pos, vel):
    """写 result2.xlsx：单时刻快照（A=标签，B=x，C=y，D=速度，224 把手）。"""
    from common import N_HANDLE

    tpl = os.path.join(ATTACH_DIR, template_name)
    out = os.path.join(ANSWER_DIR, out_name)
    wb = openpyxl.load_workbook(tpl)
    ws = wb["Sheet1"]
    for h in range(N_HANDLE):
        ws.cell(row=2 + h, column=2, value=_round6(pos[h, 0]))   # x
        ws.cell(row=2 + h, column=3, value=_round6(pos[h, 1]))   # y
        ws.cell(row=2 + h, column=4, value=_round6(vel[h]))      # 速度
    wb.save(out)
    print("  已保存: %s" % out)
    return out


def write_result3(p_min, theta_b=None):
    """写 result3.xlsx：问题3最小螺距结果（题目未给模板，生成简易表）。"""
    out = os.path.join(ANSWER_DIR, "result3.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "最小螺距"
    ws["A1"], ws["B1"] = "项目", "数值"
    ws["A2"], ws["B2"] = "最小螺距 p_min (m)", _round6(p_min)
    if theta_b is not None:
        ws["A3"], ws["B3"] = "到达边界极角 θ_b (rad)", _round6(theta_b)
        ws["A4"], ws["B4"] = "对应圈数 θ_b/(2π)", _round6(theta_b / (2 * 3.141592653589793))
    wb.save(out)
    print("  已保存: %s" % out)
    return out


def write_result5(v_max, v_max0=None, t_peak=None, h_peak=None):
    """写 result5.xlsx：问题5最大龙头速度结果（题目未给模板，生成简易表）。"""
    out = os.path.join(ANSWER_DIR, "result5.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "最大龙头速度"
    ws["A1"], ws["B1"] = "项目", "数值"
    ws["A2"], ws["B2"] = "最大龙头速度 v_max (m/s)", _round6(v_max)
    if v_max0 is not None:
        ws["A3"], ws["B3"] = "v0=1m/s 时全队最大速度 v_max^0 (m/s)", _round6(v_max0)
    if t_peak is not None:
        ws["A4"], ws["B4"] = "峰值出现时刻 t (s)", _round6(t_peak)
    if h_peak is not None:
        ws["A5"], ws["B5"] = "峰值出现把手编号", int(h_peak)
    wb.save(out)
    print("  已保存: %s" % out)
    return out


# ---- 便捷别名 ----
def write_result1(times, pos_by_t, vel_by_t):
    return write_result1_or_4("result1.xlsx", "result1.xlsx", times, pos_by_t, vel_by_t)


def write_result4(times, pos_by_t, vel_by_t):
    return write_result1_or_4("result4.xlsx", "result4.xlsx", times, pos_by_t, vel_by_t)
